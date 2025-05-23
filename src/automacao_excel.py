import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os

# Criei diretório
os.makedirs("output", exist_ok=True)

# Dados iniciais
dados = {
    "Produto": ["Notebook", "Mouse", "Teclado", "Monitor", "Cabo HDMI", "Notebook", "Mouse"],
    "Categoria": ["Informática", "Periféricos", "Periféricos", "Informática", "Acessórios", "Informática", "Periféricos"],
    "Quantidade": [5, 10, 7, 3, 15, 2, 8],
    "Valor Unitário": [3500, 50, 120, 800, 25, 3400, 55]
}
df = pd.DataFrame(dados)
df["Valor Total"] = df["Quantidade"] * df["Valor Unitário"]

# Caminho do arquivo
caminho_saida = "output/relatorio_visual.xlsx"
df.to_excel(caminho_saida, index=False)

# Abri o arquivo
wb = load_workbook(caminho_saida)
ws = wb.active
ws.title = "Relatório"

# Estilos
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)
alignment = Alignment(horizontal="center", vertical="center")

# Título "Total por Categoria"
coluna_total = 6  # Coluna F
ws.cell(row=1, column=coluna_total, value="Total por Categoria")
cell_titulo = ws.cell(row=1, column=coluna_total)
cell_titulo.font = header_font
cell_titulo.fill = header_fill
cell_titulo.border = border
cell_titulo.alignment = alignment

# Preenchi as fórmulas da coluna F
coluna_categoria = get_column_letter(2)
coluna_quantidade = get_column_letter(3)
for row in range(2, ws.max_row + 1):
    categoria_cell = f"{coluna_categoria}{row}"
    formula = f'=SUMIF({coluna_categoria}2:{coluna_categoria}{ws.max_row},{categoria_cell},{coluna_quantidade}2:{coluna_quantidade}{ws.max_row})'
    cell = ws.cell(row=row, column=coluna_total, value=formula)
    if row % 2 == 0:
        cell.fill = zebra_fill
    cell.border = border
    cell.alignment = alignment

# Apliquei estilos até coluna F
for row in range(1, ws.max_row + 1):
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        if row == 1:
            cell.font = header_font
            cell.fill = header_fill
        elif row % 2 == 0:
            cell.fill = zebra_fill
        cell.border = border
        cell.alignment = alignment

# Exemplo de Índice + Corresp (linhas 9–12, na coluna B)
linha_inicio_extras = ws.max_row + 2
ws.cell(row=linha_inicio_extras, column=1, value="Exemplo de ÍNDICE + CORRESP:")
ws.cell(row=linha_inicio_extras + 1, column=2, value='=INDEX(D2:D100, MATCH("Teclado", A2:A100, 0))')
ws.cell(row=linha_inicio_extras + 2, column=2, value='=INDEX(E2:E100, MATCH("Notebook", A2:A100, 0))')

# Apliquei bordas nas colunas A-F para linhas extras (até linha 12)
for row in range(linha_inicio_extras, linha_inicio_extras + 4):
    for col in range(1, 7):  # Colunas A-F
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = alignment

# Gráfico (abaixo de tudo)
grafico = BarChart()
grafico.title = "Total de Produtos por Categoria"
grafico.style = 10
grafico.y_axis.title = 'Quantidade'
grafico.x_axis.title = 'Categoria'

cat_range = Reference(ws, min_col=2, min_row=2, max_row=8)
val_range = Reference(ws, min_col=3, min_row=2, max_row=8)
grafico.add_data(val_range, titles_from_data=False)
grafico.set_categories(cat_range)
ws.add_chart(grafico, f"B{linha_inicio_extras + 5}")

# Ajustei largura das colunas
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

# Salvar
wb.save(caminho_saida)
print(f"✅ Arquivo final salvo em: {caminho_saida}")
